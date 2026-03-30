"""
compare_runs.py — Backtester and Excel report generator.

Replays historical BTC candle data through the strategy engine and
produces an Excel workbook with per-trade results + summary stats.

Usage:
    python compare_runs.py [--hours 24] [--output results.xlsx]
"""

import argparse
import os
import time
from datetime import datetime, timezone

from backtest import (
    fetch_historical_candles,
    group_into_windows,
    estimate_token_price,
    fetch_polymarket_history,
)
from strategy import analyze

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def run_backtest(windows, bet_amount=1.0, confidence_threshold=0.30, min_score=0.0,
                  poly_lookup=None):
    """
    Simulate the bot's strategy on historical 5-min windows.

    For each window:
      1. Use candles up to ~3 min into the window as "available data"
      2. Run strategy.analyze()
      3. Determine if bot would have bet (based on thresholds)
      4. Compare prediction to actual outcome (Polymarket resolution if available)
      5. Compute P&L using real $0.505 token pricing

    Returns list of trade dicts.
    """
    if poly_lookup is None:
        poly_lookup = {}

    trades = []

    for i, w in enumerate(windows):
        wc = w["candles"]
        if len(wc) < 3:
            continue

        # Use the latest 1-minute candle as the closest historical proxy for
        # the T-10s snipe window described by the live bot guide.
        entry_idx = len(wc) - 1
        visible_candles = wc[:entry_idx + 1]
        window_open = w["open_price"]
        current_price = visible_candles[-1]["close"]

        # Build a small lookback: previous window candles + current visible
        lookback = []
        if i > 0:
            lookback = windows[i - 1]["candles"][-15:]
        lookback.extend(visible_candles)

        score, confidence, details = analyze(
            candles=lookback,
            window_open_price=window_open,
            current_price=current_price,
            tick_prices=None,  # no tick data in backtest
        )

        prediction = "UP" if score > 0 else ("DOWN" if score < 0 else "SKIP")

        # Use Polymarket resolution if available, otherwise Binance delta
        poly_data = poly_lookup.get(w["window_ts"])
        if poly_data and poly_data.get("winner"):
            actual = poly_data["winner"].upper()
            resolution_source = "polymarket"
        else:
            actual = w["outcome"]
            resolution_source = "binance"

        if prediction == "SKIP":
            prediction = "UP" if current_price >= window_open else "DOWN"

        would_bet = True
        if prediction != "SKIP" and confidence < confidence_threshold and abs(score) >= min_score:
            would_bet = True

        # P&L using the spec's delta-based token pricing model.
        if would_bet:
            token_cost = estimate_token_price(window_open, current_price)
            correct = prediction == actual
            if correct:
                pnl = (1.0 - token_cost) * bet_amount
            else:
                pnl = -token_cost * bet_amount
        else:
            token_cost = 0
            pnl = 0
            correct = None

        ts_str = datetime.fromtimestamp(w["window_ts"], tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        trades.append({
            "window_ts": w["window_ts"],
            "time": ts_str,
            "open_price": round(window_open, 2),
            "close_price": round(w["close_price"], 2),
            "delta_pct": round((w["close_price"] - window_open) / window_open * 100, 6),
            "score": round(score, 2),
            "confidence": round(confidence, 4),
            "prediction": prediction,
            "actual": actual,
            "resolution_source": resolution_source,
            "bet": would_bet,
            "correct": correct,
            "token_cost": round(token_cost, 4),
            "pnl": round(pnl, 4),
        })

    return trades


def print_summary(trades):
    """Print a text summary of backtest results."""
    total = len(trades)
    bet_trades = [t for t in trades if t["bet"]]
    skipped = total - len(bet_trades)
    wins = [t for t in bet_trades if t["correct"]]
    losses = [t for t in bet_trades if t["correct"] is False]

    total_pnl = sum(t["pnl"] for t in bet_trades)
    avg_win = sum(t["pnl"] for t in wins) / len(wins) if wins else 0
    avg_loss = sum(t["pnl"] for t in losses) / len(losses) if losses else 0
    win_rate = len(wins) / len(bet_trades) * 100 if bet_trades else 0

    print("\n" + "=" * 60)
    print("BACKTEST SUMMARY")
    print("=" * 60)
    poly_count = sum(1 for t in trades if t.get("resolution_source") == "polymarket")
    print(f"Total 5-min windows:  {total}")
    print(f"Bets placed:          {len(bet_trades)}")
    print(f"Skipped (low conf):   {skipped}")
    print(f"Wins:                 {len(wins)}")
    print(f"Losses:               {len(losses)}")
    print(f"Win rate:             {win_rate:.1f}%")
    print(f"Total P&L:            ${total_pnl:.4f}")
    print(f"Avg win:              ${avg_win:.4f}")
    print(f"Avg loss:             ${avg_loss:.4f}")
    avg_token_cost = sum(t["token_cost"] for t in bet_trades) / len(bet_trades) if bet_trades else 0
    print(f"Avg token cost:       ${avg_token_cost:.4f}")
    print(f"Resolution source:    {poly_count} polymarket / {total - poly_count} binance")

    # Running bankroll
    if bet_trades:
        bankroll = 1.0  # starting
        peak = bankroll
        max_dd = 0
        for t in bet_trades:
            bankroll += t["pnl"]
            if bankroll > peak:
                peak = bankroll
            dd = (peak - bankroll) / peak * 100 if peak > 0 else 0
            if dd > max_dd:
                max_dd = dd
        print(f"Final bankroll:       ${bankroll:.4f} (started $1.00)")
        print(f"Max drawdown:         {max_dd:.1f}%")

    print("=" * 60)


def write_excel(trades, output_path):
    """Write backtest results to an Excel workbook."""
    if not HAS_OPENPYXL:
        print("[compare] openpyxl not installed, skipping Excel output.")
        print("  Install with: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "Time", "Open Price", "Close Price", "Delta %",
        "Score", "Confidence", "Prediction", "Actual", "Source",
        "Bet?", "Correct?", "Token Cost", "P&L", "Running P&L",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    running_pnl = 0
    for row_idx, t in enumerate(trades, 2):
        running_pnl += t["pnl"]
        values = [
            t["time"],
            t["open_price"],
            t["close_price"],
            t["delta_pct"],
            t["score"],
            t["confidence"],
            t["prediction"],
            t["actual"],
            t.get("resolution_source", "binance"),
            "YES" if t["bet"] else "no",
            "WIN" if t["correct"] is True else ("LOSS" if t["correct"] is False else "-"),
            t["token_cost"],
            t["pnl"],
            round(running_pnl, 4),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Color coding
        if t["correct"] is True:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = green_fill
        elif t["correct"] is False:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = red_fill
        elif not t["bet"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = gray_fill

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, min(len(trades) + 2, 50)):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 2

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    bet_trades = [t for t in trades if t["bet"]]
    wins = [t for t in bet_trades if t["correct"]]
    losses = [t for t in bet_trades if t["correct"] is False]
    total_pnl = sum(t["pnl"] for t in bet_trades)
    win_rate = len(wins) / len(bet_trades) * 100 if bet_trades else 0

    poly_verified = sum(1 for t in trades if t.get("resolution_source") == "polymarket")
    summary_data = [
        ("Metric", "Value"),
        ("Total Windows", len(trades)),
        ("Bets Placed", len(bet_trades)),
        ("Skipped", len(trades) - len(bet_trades)),
        ("Wins", len(wins)),
        ("Losses", len(losses)),
        ("Win Rate", f"{win_rate:.1f}%"),
        ("Total P&L", f"${total_pnl:.4f}"),
        ("Avg Win", f"${sum(t['pnl'] for t in wins) / len(wins):.4f}" if wins else "$0"),
        ("Avg Loss", f"${sum(t['pnl'] for t in losses) / len(losses):.4f}" if losses else "$0"),
        ("Avg Token Cost", f"${sum(t['token_cost'] for t in bet_trades) / len(bet_trades):.4f}" if bet_trades else "$0"),
        ("Avg Profit per Win", f"${sum((1.0 - t['token_cost']) for t in wins) / len(wins):.4f}" if wins else "$0"),
        ("Polymarket verified", f"{poly_verified}/{len(trades)}"),
    ]

    for row_idx, (metric, value) in enumerate(summary_data, 1):
        ws2.cell(row=row_idx, column=1, value=metric).font = Font(bold=(row_idx == 1))
        ws2.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            ws2.cell(row=row_idx, column=1).fill = header_fill
            ws2.cell(row=row_idx, column=1).font = header_font
            ws2.cell(row=row_idx, column=2).fill = header_fill
            ws2.cell(row=row_idx, column=2).font = header_font

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15

    # --- P&L Chart ---
    if len(bet_trades) > 1:
        ws3 = wb.create_sheet("P&L Chart Data")
        ws3.cell(row=1, column=1, value="Trade #")
        ws3.cell(row=1, column=2, value="Running P&L")
        rpnl = 0
        for i, t in enumerate(bet_trades, 1):
            rpnl += t["pnl"]
            ws3.cell(row=i + 1, column=1, value=i)
            ws3.cell(row=i + 1, column=2, value=round(rpnl, 4))

        chart = BarChart()
        chart.title = "Running P&L per Trade"
        chart.y_axis.title = "P&L ($)"
        chart.x_axis.title = "Trade #"
        chart.style = 10
        data = Reference(ws3, min_col=2, min_row=1, max_row=len(bet_trades) + 1)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=len(bet_trades) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws2.add_chart(chart, "D2")

    wb.save(output_path)
    print(f"[compare] Results saved to {output_path}")


def load_poly_csv(csv_path):
    """
    Load Polymarket resolution data from a pre-fetched CSV (e.g. poly_24h.csv).

    Returns dict: epoch_ts -> {"epoch_ts": int, "winner": str, ...}
    """
    import csv as csv_mod
    lookup = {}
    try:
        with open(csv_path, "r") as f:
            reader = csv_mod.DictReader(f)
            for row in reader:
                if row.get("resolved", "").lower() != "true":
                    continue
                epoch = int(row["epoch"])
                lookup[epoch] = {
                    "epoch_ts": epoch,
                    "winner": row.get("winner", ""),
                    "slug": row.get("slug", ""),
                    "volume": row.get("volume", ""),
                }
        print(f"[compare] Loaded {len(lookup)} resolved windows from {csv_path}")
    except Exception as e:
        print(f"[compare] Error loading CSV {csv_path}: {e}")
    return lookup


def main():
    parser = argparse.ArgumentParser(description="Backtest BTC 5-min strategy")
    parser.add_argument("--hours", type=int, default=24, help="Hours of history to test")
    parser.add_argument("--output", type=str, default="backtest_results.xlsx", help="Output Excel file")
    parser.add_argument("--bet", type=float, default=1.0, help="Bet amount per trade")
    parser.add_argument("--threshold", type=float, default=0.40, help="Min confidence threshold")
    parser.add_argument("--min-score", type=float, default=2.0, help="Min absolute score to bet")
    parser.add_argument("--poly-csv", type=str, default=None,
                        help="Pre-fetched Polymarket CSV (e.g. poly_24h.csv). "
                             "Skips live API fetch if provided.")
    args = parser.parse_args()

    print(f"[compare] Fetching {args.hours}h of BTC candle data...")
    candles = fetch_historical_candles(hours=args.hours)

    if not candles:
        print("[compare] No candles fetched. Check internet connection.")
        return

    windows = group_into_windows(candles)

    if not windows:
        print("[compare] No valid windows formed.")
        return

    # Load Polymarket resolution data
    if args.poly_csv:
        poly_lookup = load_poly_csv(args.poly_csv)
    else:
        print(f"[compare] Fetching Polymarket resolution data (may take a while)...")
        poly_markets = fetch_polymarket_history(hours=args.hours)
        poly_lookup = {pm["epoch_ts"]: pm for pm in poly_markets}
        print(f"[compare] Got {len(poly_lookup)} Polymarket-verified resolutions")

    print(f"[compare] Running strategy on {len(windows)} windows...")
    trades = run_backtest(
        windows,
        bet_amount=args.bet,
        confidence_threshold=args.threshold,
        min_score=args.min_score,
        poly_lookup=poly_lookup,
    )

    print_summary(trades)
    write_excel(trades, args.output)


if __name__ == "__main__":
    main()
